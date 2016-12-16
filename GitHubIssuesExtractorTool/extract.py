# -*- coding: utf-8 -*-

import sys
import requests
import collections
from openpyxl import Workbook

class GithubAPI:
    results = []
    raw = []
    issues_payload = {
        "per_page": 100,
        "page": 1,
        "state": "closed",
        "labels": "bug"
    }
    auth = ("theeibwen", "09196db17fc3dda1cfddb5a60f1516e5309d623b")

    def __init__(self, user, repo):
        self.repos_api_url = "https://api.github.com/repos/" + user + "/" + repo
        self.issues_api_url = self.repos_api_url + "/issues"

    def getIssues(self):
        r = requests.get(self.issues_api_url, params=self.issues_payload, auth=self.auth).json()

        while (True):
            self.raw += r

            if len(r) == 100:
                self.issues_payload["page"] += 1
                r = requests.get(self.issues_api_url, params=self.issues_payload, auth=self.auth).json()
            else:
                break

        for e in self.raw:
            print("Checking issue " + str(e["number"]))
            events = self.getEvents(e["number"])

            if len(events) > 0:
                issue = collections.OrderedDict()
                issue["id"] = e["number"]
                issue["title"] = e["title"]
                issue["description"] = e["body"]
                issue["comments"] = e["comments"]
                issue["events"] = events
                issue["created_at"] = e["created_at"]
                issue["updated_at"] = e["updated_at"]
                issue["closed_at"] = e["closed_at"]

                self.results.append(issue)

        return self.results

    def getEvents(self, issue_id):
        url = self.issues_api_url + "/" + str(issue_id) + "/events"
        raw = requests.get(url, auth=self.auth).json()
        r = []

        for result in raw:
            if not result["commit_id"]:
                continue

            commits = self.getCommit(result["commit_url"])

            if len(commits) > 0:
                event = collections.OrderedDict()

                event["id"] = result["id"]
                event["issue_id"] = issue_id
                event["commit"] = self.getCommit(result["commit_url"])
                event["created_at"] = result["created_at"]

                r.append(event)

        return r

    def getCommit(self, commit_url):
        raw = requests.get(commit_url, auth=self.auth).json()
        r = collections.OrderedDict()

        if "files" in raw.keys():
            """
            r["changes"] = raw["stats"]["total"]
            r["addition"] = raw["stats"]["additions"]
            r["deletions"] = raw["stats"]["deletions"]
            """
            r["files"] = []

            for file in raw["files"]:
                if "patch" in file.keys():
                    f = collections.OrderedDict()

                    f["filename"] = file["filename"]
                    f["patch"] = file["patch"]

                    r["files"].append(f)

        return r

if __name__ == "__main__":
    api = GithubAPI(sys.argv[1], sys.argv[2])

    print("Getting issues...")
    issues = api.getIssues()

    print ("Creating bugreports.xlsx...")
    # Creates bugreports.xlsx
    wb = Workbook()
    ws = wb.active

    headers = ["id", "title", "description", "comments", "created_at", "uploaded_at", "closed_at"]
    ws.append(headers)

    for issue in issues:
        ws.append([issue["id"], issue["title"], issue["description"].encode('utf-8'), issue["comments"], issue["created_at"],
                             issue["updated_at"], issue["closed_at"]])

    wb.save("bugreports.xlsx")

    print ("Creating events.xlsx...")
    # Creates events.xlsx
    wb = Workbook()
    ws = wb.active

    headers = ["id", "issue_id", "created_at"]
    ws.append(headers)

    for issue in issues:
        for event in issue["events"]:
            ws.append([event["id"], event["issue_id"], event["created_at"]])

    wb.save("events.xlsx")

    print ("Creating commits.xlsx...")
    # Creates commits.xlsx
    wb = Workbook()
    ws = wb.active

    i = 1
    headers = ["id", "event_id", "filename", "patch"]
    ws.append(headers)

    for issue in issues:
        for event in issue["events"]:
            for commit in event["commit"]["files"]:
                ws.append([i, event["id"], commit["filename"], commit["patch"].encode('utf-8')])

    wb.save("commits.xlsx")